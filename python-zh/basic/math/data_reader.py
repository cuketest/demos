import csv
from openpyxl import load_workbook

"""
测试数据可以通过多种方式获取，包括但不限于以下几种方法：
1. 固定的数据集合：直接在代码中定义的固定数据集合。
2. 文件读取：从外部文件（如文本文件、CSV 文件等）读取数据。
3. 数据库查询：通过数据库查询语句获取数据。
4. API 接口调用：通过调用外部 API 接口获取数据。
5. 参数化配置：使用配置文件或命令行参数传递测试数据。
6. 生成器函数：使用生成器函数生成测试数据。

根据具体情况选择适合的方式来获取测试数据，以满足测试需求。
"""


def read_data_from_csv(file_path):
    with open(file_path, newline='') as csvfile:
        data = list(csv.reader(csvfile))
    print("从CSV读取的数据：", data)
    return [(int(row[0]), int(row[1]), int(row[2])) for row in data]


def read_data_from_txt(file_path):
    with open(file_path) as txtfile:
        data = [tuple(map(int, line.split())) for line in txtfile.readlines()]
    print("从TXT读取的数据：", data)
    return data


def read_data_from_xlsx(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append((row[0], row[1], row[2]))
    print("从XLSX读取的数据：", data)
    return data
