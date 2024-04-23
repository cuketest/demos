from leanproAuto import WinAuto, Util
from pytest_html import extras
import openpyxl

# 加载Windows应用程序的UI模型文件
model = WinAuto.loadModel('model.tmodel')


# 检索应用中的表格内容
def test_retrieve_table_content(extra):
    row = 0
    column = 0
    # 输出{int}行{int}列的单元格数据
    tableModel = model.getTable("Table")
    cell = tableModel.cellValue(row, column)

    # 添加json格式数据到测试报告中
    extra.append(extras.json(cell))


    targetRow = 1
    # 读取spreadsheet中的第{int}行数据
    tableModel = model.getTable("Table")
    rowdata = tableModel.rowData(targetRow)

    # 输出所有单元格数据
    tableModel = model.getTable("Table")
    rowCount = tableModel.rowCount()
    for i in range(rowCount):
        rowData = tableModel.rowData(i)
        extra.append(extras.json(rowData))  # 输出的数据显示在报告中，请使用“运行项目”来生成报告
    attachImage(extra)

# 从文件中导入数据
def test_import_data(extra):
    fileName = "spreadsheet_data.xlsx"
    # 读取{string}文件中的数据
    worksheet = openpyxl.load_workbook("./support/data/" + fileName).worksheets[0]
    row_number = worksheet.max_row
    col_number = worksheet.max_column

    # 写入到spreadsheet中
    tableModel = model.getTable("Table")
    for row in range(1, row_number - 2):
        for col in range(col_number):
            cell = tableModel.select(row, col)
            value = worksheet.cell(row + 2, col + 1).value

            # 设置单元格值按回车
            tableModel.setCellValue(row, col, value)
            cell.pressKeys('~')
            cell.checkProperty('value', value, "写入数据不符合预期")
    attachImage(extra)

# 截屏并添加到测试报告中
def attachImage(extra):
    screenshot = model.getWindow("Window").takeScreenshot()
    extra.append(extras.image(screenshot))