from leanproAuto import WinAuto, Util
from pytest_bdd import scenarios, given, when, then, parsers
import os
import openpyxl
import pytest
import json

# 加载Windows应用程序的UI模型文件
model = WinAuto.loadModel('./models/model.tmodel')

# 加载位于"../features"目录下的所有BDD剧本文件
scenarios("../features")

""" 
- @given, @when, @then: pytest-bdd装饰器，用于定义测试的前提条件（Given）、操作步骤（When）和预期结果（Then）。
- parsers.parse: 解析器，用于解析步骤中的参数。
"""

# 检索应用中的表格内容
@given(parsers.parse('输出{row:d}行{col:d}列的单元格数据'))
def output_data(row,col, request):
    tableModel = model.getTable("Table")
    cell = tableModel.cellValue(row, col)

    # 将单元格数据添加到测试报告中
    request.attach(cell)


@when(parsers.parse('读取spreadsheet中的第{targetRow:d}行数据'))
def read_data(targetRow):
    tableModel = model.getTable("Table")
    rowdata = tableModel.rowData(targetRow)


@then('输出所有单元格数据')
def output_all_data(request):
    tableModel = model.getTable("Table")
    rowCount = tableModel.rowCount()
    for i in range(rowCount):
        rowData = tableModel.rowData(i)
        request.attach(json.dumps(rowData))
        print(rowData)


# 从文件中导入数据
# target_fixture用于将局部变量传递到上下文
@when(parsers.parse('读取{fileName}文件中的数据'), target_fixture="data")
def read_file(fileName):
    worksheet = openpyxl.load_workbook("./support/data/" + fileName).worksheets[0]
    row_number = worksheet.max_row
    col_number = worksheet.max_column
    return {"worksheet": worksheet, "row_number":row_number, "col_number":col_number}


@then("写入到spreadsheet中")
def write_file(data):
    worksheet = data['worksheet']
    row_number = data['row_number']
    col_number = data['col_number']
    tableModel = model.getTable("Table")
    for row in range(1, row_number - 2):
        for col in range(col_number):
            cell = tableModel.select(row, col)
            value = worksheet.cell(row + 2, col + 1).value
            
            # 设置单元格值按回车写入
            tableModel.setCellValue(row, col, value)
            cell.pressKeys('~')
            cell.checkProperty('value', value, "写入数据不符合预期")

