from leanproAuto import QtAuto, Util
from pytest_html import extras
import openpyxl
import csv

# 加载Qt应用程序的UI模型文件
model = QtAuto.loadModel('model.tmodel')


# 场景: 从xlsx文件中导入数据
def test_qt_table_import(extra):
    # 当读取"origin.xlsx"文件中的数据
    xlsxName = "origin.xlsx"
    worksheet = openpyxl.load_workbook("./support/" + xlsxName).worksheets[0]
    row_number = worksheet.max_row
    col_number = worksheet.max_column

    # 那么将数据写入到应用表格中
    for row in range(row_number - 1):
        for col in range(col_number):
            value = worksheet.cell(row + 2, col + 1).value
            model.getTable("Table").setCellValue(row, col, value)
    attachImage(extra)


# 场景: 从应用中导出数据到其它应用中
def test_qt_table_export(extra):
    # 当读取表格中全部数据
    data = model.getTable("Table").data()  # 表格数据
    header = model.getTable("Table").columnHeaders()  # 表头（每列）
    rowHeaders = model.getTable("Table").rowHeaders()  # 表头（每行）

    # 将数据转换为 JSON 格式并添加到测试报告中
    extra.append(extras.json(data))
    extra.append(extras.json(header))
    extra.append(extras.json(rowHeaders))

    # 那么将数据写入到xlsx文件"data.xlsx"中
    xlsxName = "data.xlsx"
    filename = "./support/" + xlsxName
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet('Sheet1', 0)
    worksheet.append(header)
    for row in data:
        worksheet.append(row)
    workbook.save(filename)

    # 那么将数据写入到CSV文件"data.csv"中
    csvName = "data.csv"
    filePath = "./support/" + csvName
    csvFile = open(filePath, "w", newline="")
    writer = csv.writer(csvFile)
    writer.writerow(header)
    writer.writerows(data)
    attachImage(extra)


# 场景: 单元格操作
def test_qt_table_cell_manipulation(extra):
    # 假如目标单元格在第90行第0列
    row = 90
    col = 0
    targetCell = model.getTable("Table").getItem(row, col)
    value = targetCell.value()
    extra.append(extras.text(value))

    # 当修改数据为"New Value!"并验证
    value = "New Value!"
    extra.append(extras.text(targetCell.value()))
    targetCell.set(value)

    # 检查控件value属性是否为期待值,否则输出指定信息
    targetCell.checkProperty("value", value, '修改后的值不为' + repr(value))

    # 那么滚动到目标单元格
    targetCell.scrollIntoView()
    attachImage(extra)

# 截屏并添加到测试报告中
def attachImage(extra):
    screenshot = model.getWindow("Dock_Widgets").takeScreenshot()
    extra.append(extras.image(screenshot))