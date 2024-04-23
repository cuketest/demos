from leanproAuto import QtAuto, Util
from pytest_bdd import scenarios, scenario, given, when, then, parsers
import openpyxl
import csv

# 加载Qt应用程序的UI模型文件
model = QtAuto.loadModel('models/model.tmodel')

# 加载位于"../features"目录下的所有BDD剧本文件
scenarios("../features")

""" 
- @given, @when, @then: pytest-bdd装饰器，用于定义测试的前提条件（Given）、操作步骤（When）和预期结果（Then）。
- parsers.parse: 解析器，用于解析步骤中的参数。
"""

# 场景: 从xlsx文件中导入数据
@when(parsers.parse('读取"{xlsxName}"文件中的数据'), target_fixture = "number")
def read_file():
    xlsxName = "origin.xlsx"
    worksheet = openpyxl.load_workbook("./support/" + xlsxName).worksheets[0]
    row_number = worksheet.max_row
    col_number = worksheet.max_column
    return{'worksheet':worksheet,'row_number':row_number,'col_number':col_number}


@then('将数据写入到应用表格中')
def write_table(number):
    worksheet = number['worksheet']
    row_number = number['row_number']
    col_number = number['col_number']
    for row in range(row_number - 1):
        for col in range(col_number):
            value = worksheet.cell(row + 2, col + 1).value
            model.getTable("Table").setCellValue(row, col, value)
    screenshot = model.getWindow("Dock_Widgets").takeScreenshot()



# 场景: 从应用中导出数据到其它应用中
@when(parsers.parse('读取表格中全部数据'), target_fixture = "table")
def read_table():
    data = model.getTable("Table").data()  # 表格数据
    header = model.getTable("Table").columnHeaders()  # 表头（每列）
    rowHeaders = model.getTable("Table").rowHeaders()  # 表头（每行）
    return {'data':data, 'header':header}


@then(parsers.parse('将数据写入到xlsx文件"{xlsxName}"中'))
def write_xlsx(xlsxName,table):
    data = table['data']
    header = table['header']
    filename = "./support/" + xlsxName

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 在工作簿中创建一个名为 'Sheet1' 的工作表，并将其置于索引 0 的位置
    worksheet = workbook.create_sheet('Sheet1', 0)

    # 将表头添加到工作表中
    worksheet.append(header)

    # 将数据逐行添加到工作表中
    for row in data:
        worksheet.append(row)

    # 保存工作簿到指定的文件路径
    workbook.save(filename)

@then(parsers.parse('将数据写入到CSV文件"{csvName}"中'))
def write_csv(csvName, table):
    data = table['data']
    header = table['header']
    filePath = "./support/" + csvName

    # 打开 CSV 文件进行写入
    csvFile = open(filePath, "w", newline="")
    writer = csv.writer(csvFile)
    writer.writerow(header)
    writer.writerows(data)


# 场景: 单元格操作
@given(parsers.parse('目标单元格在第{row:d}行第{col:d}列'), target_fixture = "targetCell")
def target_cell(row, col, request):
    targetCell = model.getTable("Table").getItem(row, col)
    value = targetCell.value()

    # 将单元格值添加到测试报告中
    request.attach(value)
    return targetCell


@when(parsers.parse('修改数据为"{value}"并验证'))
def modify_value(value,targetCell, request):
    request.attach(targetCell.value())
    print(targetCell.value())
    targetCell.set(value)

    # 检查目标单元格的值是否等于指定值，否则输出指定信息
    targetCell.checkProperty("value", value, '修改后的值不为' + repr(value))


@then('滚动到目标单元格')
def scroll_into_view(targetCell):
    targetCell.scrollIntoView()
